"""
–ú–æ–¥—É–ª—å –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞ –∏—Å—Ç–æ—Ä–∏–∏ —Å–¥–µ–ª–æ–∫ –≤ Excel –¥–ª—è –∞–Ω–∞–ª–∏–∑–∞.
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Optional
import logging

from bot.state import TradeRecord

logger = logging.getLogger(__name__)


def export_trades_to_excel(
    trades: List[TradeRecord],
    output_dir: str = "trade_history",
    filename: Optional[str] = None
) -> str:
    """
    –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç –∏—Å—Ç–æ—Ä–∏—é —Å–¥–µ–ª–æ–∫ –≤ Excel —Ñ–∞–π–ª.
    
    Args:
        trades: –°–ø–∏—Å–æ–∫ —Å–¥–µ–ª–æ–∫ –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞
        output_dir: –î–∏—Ä–µ–∫—Ç–æ—Ä–∏—è –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ñ–∞–π–ª–∞
        filename: –ò–º—è —Ñ–∞–π–ª–∞ (–µ—Å–ª–∏ –Ω–µ —É–∫–∞–∑–∞–Ω–æ, –≥–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏)
    
    Returns:
        –ü—É—Ç—å –∫ —Å–æ–∑–¥–∞–Ω–Ω–æ–º—É Excel —Ñ–∞–π–ª—É
    """
    if not trades:
        logger.warning("No trades to export")
        return ""
    
    logger.info(f"üìä Starting export of {len(trades)} trade(s) to Excel...")
    
    # –°–æ–∑–¥–∞–µ–º –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏—é –µ—Å–ª–∏ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    logger.debug(f"Output directory: {output_path}")
    
    # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –∏–º—è —Ñ–∞–π–ª–∞ –µ—Å–ª–∏ –Ω–µ —É–∫–∞–∑–∞–Ω–æ
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"trades_history_{timestamp}.xlsx"
    
    filepath = output_path / filename
    
    # –ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞
    data = []
    for trade in trades:
        # –ò–∑–≤–ª–µ–∫–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –º–æ–¥–µ–ª–∏
        model_name = trade.model_name or ""
        model_type = ""
        if model_name:
            # –ü—ã—Ç–∞–µ–º—Å—è –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å —Ç–∏–ø –º–æ–¥–µ–ª–∏ –∏–∑ –∏–º–µ–Ω–∏ —Ñ–∞–π–ª–∞
            if "ensemble" in model_name.lower():
                if "triple" in model_name.lower():
                    model_type = "TripleEnsemble"
                elif "quad" in model_name.lower():
                    model_type = "QuadEnsemble"
                else:
                    model_type = "Ensemble"
            elif "rf_" in model_name.lower() or "random" in model_name.lower():
                model_type = "RandomForest"
            elif "xgb_" in model_name.lower() or "xgboost" in model_name.lower():
                model_type = "XGBoost"
            elif "lgb_" in model_name.lower() or "lightgbm" in model_name.lower():
                model_type = "LightGBM"
            else:
                model_type = "Unknown"
        
        # –ò–∑–≤–ª–µ–∫–∞–µ–º –ø–∞—Ä–∞–º–µ—Ç—Ä—ã –∏–∑ signal_parameters
        signal_params = trade.signal_parameters or {}
        tp_pct = signal_params.get('take_profit_pct', None)
        sl_pct = signal_params.get('stop_loss_pct', None)
        rr_ratio = signal_params.get('risk_reward_ratio', None)
        
        # –í—ã—á–∏—Å–ª—è–µ–º –¥–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å —Å–¥–µ–ª–∫–∏
        duration_minutes = None
        if trade.entry_time and trade.exit_time:
            try:
                entry_dt = datetime.fromisoformat(trade.entry_time.replace('Z', '+00:00'))
                exit_dt = datetime.fromisoformat(trade.exit_time.replace('Z', '+00:00'))
                duration = exit_dt - entry_dt
                duration_minutes = duration.total_seconds() / 60
            except Exception as e:
                logger.debug(f"Could not calculate duration: {e}")
        
        # –í—ã—á–∏—Å–ª—è–µ–º —Ä–∞–∑–º–µ—Ä –ø–æ–∑–∏—Ü–∏–∏ –≤ USD
        position_size_usd = trade.entry_price * trade.qty if trade.entry_price and trade.qty else 0.0
        
        # –§–æ—Ä–º–∏—Ä—É–µ–º —Å—Ç—Ä–æ–∫—É –¥–∞–Ω–Ω—ã—Ö
        row = {
            # –û—Å–Ω–æ–≤–Ω–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è
            "–í—Ä–µ–º—è –≤—Ö–æ–¥–∞": trade.entry_time,
            "–í—Ä–µ–º—è –≤—ã—Ö–æ–¥–∞": trade.exit_time or "",
            "–î–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å (–º–∏–Ω)": duration_minutes,
            "–°–∏–º–≤–æ–ª": trade.symbol,
            "–ù–∞–ø—Ä–∞–≤–ª–µ–Ω–∏–µ": trade.side,
            "–°—Ç–∞—Ç—É—Å": trade.status,
            
            # –¶–µ–Ω—ã
            "–¶–µ–Ω–∞ –≤—Ö–æ–¥–∞": trade.entry_price,
            "–¶–µ–Ω–∞ –≤—ã—Ö–æ–¥–∞": trade.exit_price or "",
            "Take Profit": trade.take_profit or "",
            "Stop Loss": trade.stop_loss or "",
            
            # –†–∞–∑–º–µ—Ä—ã –∏ PnL
            "–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ": trade.qty,
            "–†–∞–∑–º–µ—Ä –ø–æ–∑–∏—Ü–∏–∏ (USD)": position_size_usd,
            "–ú–∞—Ä–∂–∞ (USD)": trade.margin_usd,
            "–ü–ª–µ—á–æ": trade.leverage,
            "PnL (USD)": trade.pnl_usd,
            "PnL (%)": trade.pnl_pct,
            
            # ML –º–æ–¥–µ–ª—å
            "–ú–æ–¥–µ–ª—å ML": model_name,
            "–¢–∏–ø –º–æ–¥–µ–ª–∏": model_type,
            "–ì–æ—Ä–∏–∑–æ–Ω—Ç": trade.horizon,
            
            # –°–∏–≥–Ω–∞–ª
            "–ü—Ä–∏—á–∏–Ω–∞ –≤—Ö–æ–¥–∞": trade.entry_reason,
            "–ü—Ä–∏—á–∏–Ω–∞ –≤—ã—Ö–æ–¥–∞": trade.exit_reason,
            "–£–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å (%)": trade.confidence * 100 if trade.confidence else 0,
            "–°–∏–ª–∞ —Å–∏–≥–Ω–∞–ª–∞": trade.signal_strength,
            
            # –ü–∞—Ä–∞–º–µ—Ç—Ä—ã —Å–∏–≥–Ω–∞–ª–∞
            "TP (%)": tp_pct * 100 if tp_pct else "",
            "SL (%)": sl_pct * 100 if sl_pct else "",
            "Risk/Reward": rr_ratio,
            
            # –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ
            "DCA —Å—á–µ—Ç—á–∏–∫": trade.dca_count,
        }
        
        data.append(row)
    
    # –°–æ–∑–¥–∞–µ–º DataFrame
    df = pd.DataFrame(data)
    
    # –°–æ—Ä—Ç–∏—Ä—É–µ–º –ø–æ –≤—Ä–µ–º–µ–Ω–∏ –≤—Ö–æ–¥–∞ (–Ω–æ–≤—ã–µ —Å–≤–µ—Ä—Ö—É)
    if "–í—Ä–µ–º—è –≤—Ö–æ–¥–∞" in df.columns:
        df = df.sort_values("–í—Ä–µ–º—è –≤—Ö–æ–¥–∞", ascending=False)
    
    try:
        # –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ–º –≤ Excel —Å —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='–°–¥–µ–ª–∫–∏', index=False)
            
            # –ü–æ–ª—É—á–∞–µ–º workbook –∏ worksheet –¥–ª—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
            workbook = writer.book
            worksheet = writer.sheets['–°–¥–µ–ª–∫–∏']
            
            # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∞—è —à–∏—Ä–∏–Ω–∞ –∫–æ–ª–æ–Ω–æ–∫
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                )
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —á–∏—Å–ª–æ–≤—ã—Ö –∫–æ–ª–æ–Ω–æ–∫
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00
            
            # –ó–∞–≥–æ–ª–æ–≤–∫–∏
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ PnL - –∑–µ–ª–µ–Ω—ã–π –¥–ª—è –ø—Ä–∏–±—ã–ª–∏, –∫—Ä–∞—Å–Ω—ã–π –¥–ª—è —É–±—ã—Ç–∫–∞
            pnl_col_idx = None
            for idx, col in enumerate(df.columns, 1):
                if col == "PnL (USD)":
                    pnl_col_idx = idx
                    break
            
            if pnl_col_idx:
                pnl_fill_profit = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                pnl_fill_loss = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=pnl_col_idx)
                    pnl_value = df.iloc[row_idx - 2]["PnL (USD)"]
                    if isinstance(pnl_value, (int, float)):
                        if pnl_value > 0:
                            cell.fill = pnl_fill_profit
                        elif pnl_value < 0:
                            cell.fill = pnl_fill_loss
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –ø—Ä–æ—Ü–µ–Ω—Ç–æ–≤
            for idx, col in enumerate(df.columns, 1):
                if "%" in col or col == "–£–≤–µ—Ä–µ–Ω–Ω–æ—Å—Ç—å (%)":
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_idx, column=idx)
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = FORMAT_PERCENTAGE_00
        
        logger.info(f"‚úÖ –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {len(trades)} —Å–¥–µ–ª–æ–∫ –≤ {filepath}")
        return str(filepath)
    
    except ImportError:
        # –ï—Å–ª–∏ openpyxl –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω, —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ–º –±–µ–∑ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
        logger.warning("openpyxl not installed, exporting without formatting")
        df.to_excel(filepath, index=False)
        return str(filepath)
    except Exception as e:
        logger.error(f"Error exporting trades to Excel: {e}", exc_info=True)
        # –ü—Ä–æ–±—É–µ–º —ç–∫—Å–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å –≤ CSV –∫–∞–∫ fallback
        csv_path = filepath.with_suffix('.csv')
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        logger.info(f"Exported to CSV instead: {csv_path}")
        return str(csv_path)


def export_closed_trades_to_excel(
    trades: List[TradeRecord],
    output_dir: str = "trade_history",
    filename: Optional[str] = None
) -> str:
    """
    –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç —Ç–æ–ª—å–∫–æ –∑–∞–∫—Ä—ã—Ç—ã–µ —Å–¥–µ–ª–∫–∏ –≤ Excel —Ñ–∞–π–ª.
    
    Args:
        trades: –°–ø–∏—Å–æ–∫ –≤—Å–µ—Ö —Å–¥–µ–ª–æ–∫
        output_dir: –î–∏—Ä–µ–∫—Ç–æ—Ä–∏—è –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ñ–∞–π–ª–∞
        filename: –ò–º—è —Ñ–∞–π–ª–∞ (–µ—Å–ª–∏ –Ω–µ —É–∫–∞–∑–∞–Ω–æ, –≥–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏)
    
    Returns:
        –ü—É—Ç—å –∫ —Å–æ–∑–¥–∞–Ω–Ω–æ–º—É Excel —Ñ–∞–π–ª—É
    """
    closed_trades = [t for t in trades if t.status == "closed"]
    return export_trades_to_excel(closed_trades, output_dir, filename)
